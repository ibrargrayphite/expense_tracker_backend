from rest_framework import viewsets, permissions, mixins
from rest_framework.validators import ValidationError
from rest_framework import status
from django.db import transaction
from django.http import HttpResponse
from tracker.models import InternalTransaction, Transaction, TransactionAccount, TransactionSplit, Loan, Account
from tracker.serializers.transaction import InternalTransactionSerializer, TransactionSerializer
from tracker.pagination import TransactionResultsSetPagination
import io
import logging
import requests
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from django.db.models import Sum
from django.conf import settings
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from tracker.filters import TransactionFilter, InternalTransactionFilter
from rest_framework.response import Response
from rest_framework.decorators import action
from drf_spectacular.utils import (
    extend_schema, extend_schema_view, OpenApiParameter, OpenApiExample,
    OpenApiResponse, inline_serializer
)
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers as drf_serializers

logger = logging.getLogger(__name__)

@extend_schema_view(
    list=extend_schema(
        tags=["Transactions"],
        summary="List transactions",
        description=(
            "Returns a paginated list of all transactions for the authenticated user, ordered newest first.\n\n"
            "Each transaction can be one of:\n"
            "- **Standard** – INCOME or EXPENSE with splits\n"
            "- **Loan-related** – LOAN_TAKEN, MONEY_LENT, LOAN_REPAYMENT, or REIMBURSEMENT\n"
            "- **Internal Transfer** – auto-serialized via `InternalTransactionSerializer` with `is_internal=true`\n\n"
            "**Filterable by:** `type`, `account`, `contact`, `expense_category`, `income_source`, `start_date`, `end_date`, `min_amount`, `max_amount`\n\n"
            "**Searchable by:** contact name, split notes, account name\n\n"
            "**Sortable by:** `date`, `created_at`, `amount`"
        ),
        parameters=[
            OpenApiParameter(
                "type", OpenApiTypes.STR,
                description=(
                    "Filter by transaction type. "
                    "Choices: `INCOME`, `EXPENSE`, `LOAN_TAKEN`, `MONEY_LENT`, `LOAN_REPAYMENT`, `REIMBURSEMENT`, `TRANSFER`."
                )
            ),
            OpenApiParameter("account", OpenApiTypes.INT, description="Filter by account ID (user's bank account)."),
            OpenApiParameter("contact", OpenApiTypes.INT, description="Filter by contact ID."),
            OpenApiParameter("expense_category", OpenApiTypes.INT, description="Filter by expense category ID."),
            OpenApiParameter("income_source", OpenApiTypes.INT, description="Filter by income source ID."),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Filter transactions on or after this date (YYYY-MM-DD)."),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Filter transactions on or before this date (YYYY-MM-DD)."),
            OpenApiParameter("min_amount", OpenApiTypes.NUMBER, description="Filter transactions with total amount ≥ this value."),
            OpenApiParameter("max_amount", OpenApiTypes.NUMBER, description="Filter transactions with total amount ≤ this value."),
            OpenApiParameter("search", OpenApiTypes.STR, description="Full-text search across contact names, split notes, and internal transfer notes."),
            OpenApiParameter("ordering", OpenApiTypes.STR, description="Sort field. Prefix with `-` for descending. Options: `date`, `created_at`, `amount`."),
        ],
        responses={200: TransactionSerializer(many=True)},
    ),
    create=extend_schema(
        tags=["Transactions"],
        summary="Create a transaction",
        description=(
            "Creates a new income, expense, or loan-related transaction.\n\n"
            "**Payload structure:**\n"
            "```json\n"
            "{\n"
            "  \"date\": \"2024-01-15T10:30:00Z\",\n"
            "  \"contact\": <contact_id | null>,\n"
            "  \"contact_account\": <contact_account_id | null>,\n"
            "  \"accounts\": [\n"
            "    {\n"
            "      \"account\": <user_account_id>,\n"
            "      \"splits\": [\n"
            "        {\n"
            "          \"type\": \"INCOME|EXPENSE|LOAN_TAKEN|MONEY_LENT|LOAN_REPAYMENT|REIMBURSEMENT\",\n"
            "          \"amount\": \"500.00\",\n"
            "          \"income_source\": <id | null>,\n"
            "          \"expense_category\": <id | null>,\n"
            "          \"loan\": <id | null>,\n"
            "          \"note\": \"optional note\"\n"
            "        }\n"
            "      ]\n"
            "    }\n"
            "  ]\n"
            "}\n"
            "```\n\n"
            "**Type-specific rules:**\n\n"
            "| Type | contact | contact_account | income_source | expense_category | loan |\n"
            "|------|---------|-----------------|---------------|------------------|------|\n"
            "| `INCOME` | ❌ | ❌ | ✅ Required | ❌ | ❌ |\n"
            "| `EXPENSE` | ❌ | ❌ | ❌ | ✅ Required | ❌ |\n"
            "| `LOAN_TAKEN` | ✅ Required | ✅ Required | ❌ | ❌ | ❌ (auto-created) |\n"
            "| `MONEY_LENT` | ✅ Required | ✅ Required | ❌ | ❌ | ❌ (auto-created) |\n"
            "| `LOAN_REPAYMENT` | ✅ Required | ✅ Required | ❌ | ❌ | ✅ Required |\n"
            "| `REIMBURSEMENT` | ✅ Required | ✅ Required | ❌ | ❌ | ✅ Required |\n\n"
            "**Balance rules:**\n"
            "- `EXPENSE`, `MONEY_LENT`, `LOAN_REPAYMENT` deduct from account balance (balance must be sufficient).\n"
            "- `INCOME`, `LOAN_TAKEN`, `REIMBURSEMENT` add to account balance."
        ),
        request=TransactionSerializer,
        responses={
            201: TransactionSerializer,
            400: OpenApiResponse(description="Validation error (missing required fields, insufficient balance, invalid loan, etc.)."),
        },
        examples=[
            OpenApiExample(
                "EXPENSE transaction",
                request_only=True,
                value={
                    "date": "2024-01-15T10:30:00Z",
                    "accounts": [
                        {
                            "account": 1,
                            "splits": [
                                {"type": "EXPENSE", "amount": "150.00", "expense_category": 2, "note": "Groceries at supermarket"}
                            ]
                        }
                    ]
                },
            ),
            OpenApiExample(
                "INCOME transaction",
                request_only=True,
                value={
                    "date": "2024-01-15T09:00:00Z",
                    "accounts": [
                        {
                            "account": 1,
                            "splits": [
                                {"type": "INCOME", "amount": "5000.00", "income_source": 1, "note": "Monthly salary"}
                            ]
                        }
                    ]
                },
            ),
            OpenApiExample(
                "LOAN_REPAYMENT transaction",
                request_only=True,
                value={
                    "date": "2024-01-15T12:00:00Z",
                    "contact": 3,
                    "contact_account": 5,
                    "accounts": [
                        {
                            "account": 1,
                            "splits": [
                                {"type": "LOAN_REPAYMENT", "amount": "200.00", "loan": 2, "note": "Partial repayment"}
                            ]
                        }
                    ]
                },
            ),
        ],
    ),
    retrieve=extend_schema(
        tags=["Transactions"],
        summary="Retrieve a transaction",
        description="Returns full details of a single transaction including all accounts and splits.",
        responses={200: TransactionSerializer},
    ),
)
class TransactionViewSet(mixins.CreateModelMixin,
                         mixins.ListModelMixin,
                         mixins.RetrieveModelMixin,
                         viewsets.GenericViewSet):
    """
    Create and Read for transactions.

    list          GET    /api/transactions/
    create        POST   /api/transactions/
    retrieve      GET    /api/transactions/{id}/
    upload_image  PATCH  /api/transactions/{id}/upload_image/
    """
    serializer_class = TransactionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = TransactionResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = TransactionFilter
    search_fields = ['contact__first_name', 'contact__last_name', 'accounts__splits__note', 'accounts__account__account_name', 'accounts__account__bank_name']
    ordering_fields = ['date', 'created_at', 'amount']

    @extend_schema(
        tags=["Transactions"],
        summary="Upload a receipt image",
        description=(
            "Uploads or replaces the receipt/image attached to an existing transaction.\n\n"
            "Send as **multipart/form-data** with a single `image` field containing the image file.\n\n"
            "Supported formats: JPEG, PNG, WebP, GIF."
        ),
        request=inline_serializer(
            name="UploadImageRequest",
            fields={"image": drf_serializers.ImageField(help_text="Image file (JPEG, PNG, WebP, GIF).")}
        ),
        responses={
            200: OpenApiResponse(
                description="Image uploaded successfully.",
                examples=[OpenApiExample("Success", value={"detail": "Image uploaded successfully."})],
            ),
            400: OpenApiResponse(
                description="No image file provided.",
                examples=[OpenApiExample("Missing image", value={"image": "No image file provided."})],
            ),
        },
    )
    @action(detail=True, methods=['patch'], url_path='upload_image')
    def upload_image(self, request, pk=None):
        """Patch only the image field of an existing transaction."""
        instance = self.get_object()
        image = request.FILES.get('image')
        if not image:
            return Response({'image': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        instance.image = image
        instance.save(update_fields=['image'])
        return Response({'detail': 'Image uploaded successfully.'}, status=status.HTTP_200_OK)

    @extend_schema(
        tags=["Transactions"],
        summary="Export transactions to Excel",
        description=(
            "Generates and downloads an Excel (.xlsx) file containing all transactions (both regular and internal transfers).\n\n"
            "Columns: `Date`, `From`, `To`, `Type`, `Amount`, `Category`, `Note`, `Receipt`\n\n"
            "**Optional date filters:**\n"
            "- `start_date` – Include transactions from this date (YYYY-MM-DD)\n"
            "- `end_date` – Include transactions up to this date (YYYY-MM-DD)\n\n"
            "Returns a file download with `Content-Disposition: attachment; filename=transactions.xlsx`."
        ),
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Export from this date (YYYY-MM-DD)."),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Export up to this date (YYYY-MM-DD)."),
        ],
        responses={
            200: OpenApiResponse(description="Excel file download."),
        },
    )
    async def _fetch_transaction_image(self, client, url):
        """Helper to fetch a single image asynchronously."""
        try:
            resp = await client.get(url, timeout=10)
            resp.raise_for_status()
            return io.BytesIO(resp.content)
        except Exception as e:
            logger.error("Failed to fetch image from %s: %s", url, e)
            return None

    @extend_schema(
        tags=["Transactions"],
        summary="Export transactions to Excel",
        description=(
            "Generates an Excel (`.xlsx`) file containing all transactions and internal transfers.\n\n"
            "**Async Features:**\n"
            "- Parallel image fetching from Cloudinary (production).\n"
            "- Non-blocking event loop during I/O.\n\n"
            "**Optional Filters:**\n"
            "- `start_date` – Include transactions from this date (YYYY-MM-DD)\n"
            "- `end_date` – Include transactions up to this date (YYYY-MM-DD)\n\n"
            "Returns a file download with `Content-Disposition: attachment; filename=transactions.xlsx`."
        ),
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Export from this date (YYYY-MM-DD)."),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Export up to this date (YYYY-MM-DD)."),
        ],
        responses={
            200: OpenApiResponse(description="Excel file download."),
        },
    )
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        from asgiref.sync import async_to_sync
        return async_to_sync(self._export_excel_async)(request)

    async def _export_excel_async(self, request):
        from asgiref.sync import sync_to_async
        import httpx
        import asyncio
        
        user = request.user
        
        # Get date filter params if any
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Fetch Transactions (wrapped in sync_to_async)
        def get_transactions():
            qs = Transaction.objects.filter(user=user).prefetch_related(
                'accounts__account', 
                'accounts__splits',
                'contact',
                'contact_account'
            )
            if start_date: qs = qs.filter(date__gte=start_date)
            if end_date: qs = qs.filter(date__lte=end_date)
            return list(qs.order_by('-date', '-created_at'))

        def get_internal_transactions():
            qs = InternalTransaction.objects.filter(user=user).select_related('from_account', 'to_account')
            if start_date: qs = qs.filter(date__gte=start_date)
            if end_date: qs = qs.filter(date__lte=end_date)
            return list(qs.order_by('-date', '-created_at'))

        tx_list = await sync_to_async(get_transactions)()
        it_list = await sync_to_async(get_internal_transactions)()

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Define Headers
        headers = ["Date", "From", "To", "Type", "Amount", "Category", "Note", "Receipt"]
        ws.append(headers)
        RECEIPT_COL = len(headers)

        # Style Headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Combine and Sort
        all_data = []
        
        # Internal Transactions
        for it in it_list:
            all_data.append({
                'date': it.date,
                'created_at': it.created_at,
                'from': f"{it.from_account.account_name} - {it.from_account.bank_name} - {it.from_account.account_number}",
                'to': f"{it.to_account.account_name} - {it.to_account.bank_name} - {it.to_account.account_number}",
                'type': "Transfer",
                'amount': float(it.amount),
                'category': "-",
                'note': it.note or "-",
                'image_url': None,
                'image_field': None,
            })

        # Regular Transactions
        # We need to calculate amount via aggregate - let's do it in sync helper
        def get_tx_data(tx):
            total_amt = tx.accounts.aggregate(total=Sum('splits__amount'))['total'] or 0
            first_split = None
            for acc in tx.accounts.all():
                if acc.splits.exists():
                    first_split = acc.splits.first()
                    break
            
            tx_type = first_split.type if first_split else "-"
            category = "-"
            if first_split:
                if first_split.expense_category: category = first_split.expense_category.name
                elif first_split.income_source: category = first_split.income_source.name

            user_accounts = ", ".join([f"{acc.account.account_name} - {acc.account.bank_name} - {acc.account.account_number}" for acc in tx.accounts.all()])
            contact_acc = "-"
            if tx.contact_account:
                contact_acc = f"{tx.contact_account.account_name} - {tx.contact_account.bank_name} - {tx.contact_account.account_number}"
            elif tx.contact:
                contact_acc = f"{tx.contact.first_name} {tx.contact.last_name}"

            if tx_type == 'EXPENSE':
                from_display, to_display = user_accounts, "-"
            elif tx_type == 'INCOME':
                from_display, to_display = "-", user_accounts
            elif tx_type in ['LOAN_TAKEN', 'REIMBURSEMENT']:
                from_display, to_display = contact_acc, user_accounts
            elif tx_type in ['LOAN_REPAYMENT', 'MONEY_LENT']:
                from_display, to_display = user_accounts, contact_acc
            else:
                from_display, to_display = "-", "-"

            return {
                'date': tx.date,
                'created_at': tx.created_at,
                'from': from_display,
                'to': to_display,
                'type': tx_type.replace('_', ' ').title(),
                'amount': float(total_amt),
                'category': category,
                'note': first_split.note if first_split else "-",
                'image_url': tx.image.url if tx.image else None,
                'image_field': tx.image if tx.image else None,
            }

        for tx in tx_list:
            all_data.append(await sync_to_async(get_tx_data)(tx))

        # Sort newest first
        all_data.sort(key=lambda x: (x['date'], x['created_at']), reverse=True)

        # Parallel Image Fetching
        image_bytes_map = {}
        is_prod = getattr(settings, 'ENVIRONMENT', 'development') == 'production'
        
        if any(d['image_url'] or d['image_field'] for d in all_data):
            async with httpx.AsyncClient() as client:
                fetch_tasks = []
                # Map task result back to index
                task_to_idx = []
                
                for idx, row in enumerate(all_data):
                    if is_prod and row['image_url']:
                        fetch_tasks.append(self._fetch_transaction_image(client, row['image_url']))
                        task_to_idx.append(idx)
                    elif not is_prod and row['image_field']:
                        # Local file read is sync but blocking; we wrap it
                        fetch_tasks.append(sync_to_async(lambda field: io.BytesIO(field.read()))(row['image_field']))
                        task_to_idx.append(idx)
                
                if fetch_tasks:
                    results = await asyncio.gather(*fetch_tasks)
                    for idx, img_bytes in zip(task_to_idx, results):
                        image_bytes_map[idx] = img_bytes

        # Append rows and embed images
        IMG_WIDTH, IMG_HEIGHT, ROW_HEIGHT, COL_WIDTH = 120, 90, 75, 20
        
        for idx, row_data in enumerate(all_data):
            excel_row = ws.max_row + 1
            ws.append([
                row_data['date'].strftime("%Y-%m-%d %H:%M"),
                row_data['from'],
                row_data['to'],
                row_data['type'],
                row_data['amount'],
                row_data['category'],
                row_data['note'],
                None,
            ])

            img_bytes = image_bytes_map.get(idx)
            if img_bytes:
                try:
                    img_bytes.seek(0)
                    xl_img = XLImage(img_bytes)
                    xl_img.width, xl_img.height = IMG_WIDTH, IMG_HEIGHT
                    from openpyxl.utils import get_column_letter
                    cell_anchor = f"{get_column_letter(RECEIPT_COL)}{excel_row}"
                    ws.add_image(xl_img, cell_anchor)
                    ws.row_dimensions[excel_row].height = ROW_HEIGHT
                except Exception as e:
                    logger.error("Failed to embed image in Excel for row %s: %s", excel_row, e)

        # Auto-adjust columns
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(RECEIPT_COL)].width = COL_WIDTH
        for column in ws.columns:
            col_letter = column[0].column_letter
            if col_letter == get_column_letter(RECEIPT_COL): continue
            max_length = max((len(str(cell.value or "")) for cell in column), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
        # wb.save is blocking; wrap it
        await sync_to_async(wb.save)(response)
        return response

    def get_queryset(self):
        from django.db.models.functions import Coalesce
        from django.db.models import DecimalField, Prefetch
        from decimal import Decimal
        return Transaction.objects.filter(
            user=self.request.user
        ).select_related(
            'contact',
            'contact_account',
            'internal_transaction',
            'internal_transaction__from_account',
            'internal_transaction__to_account',
        ).prefetch_related(
            Prefetch(
                'accounts',
                queryset=TransactionAccount.objects.select_related('account').prefetch_related(
                    Prefetch(
                        'splits',
                        queryset=TransactionSplit.objects.select_related(
                            'expense_category', 'income_source', 'loan', 'loan__contact'
                        )
                    )
                )
            )
        ).annotate(
            amount=Coalesce(
                Sum('accounts__splits__amount'),
                'internal_transaction__amount',
                Decimal('0.0'),
                output_field=DecimalField()
            )
        ).order_by('-date', '-created_at')

    @transaction.atomic
    def perform_create(self, serializer):
        try:
            serializer.save()
            logger.info("Transaction created by user %s", self.request.user.id)
        except ValidationError as e:
            logger.warning("Transaction validation failed for user %s: %s", self.request.user.id, e)
            raise e
        except Exception as e:
            logger.error("Transaction creation error for user %s: %s", self.request.user.id, e, exc_info=True)
            raise ValidationError({'detail': str(e)})

    def list(self, request, *args, **kwargs):
        from django.core.cache import cache
        from tracker.cache import transactions_list_key, CACHE_TTL

        _FILTER_PARAMS = {
            'type', 'account', 'contact', 'expense_category',
            'income_source', 'start_date', 'end_date',
            'min_amount', 'max_amount', 'search', 'ordering', 'page',
        }
        has_filters = any(request.query_params.get(k) for k in _FILTER_PARAMS)

        if not has_filters:
            cache_key = transactions_list_key(request.user.id)
            cached = cache.get(cache_key)
            if cached is not None:
                logger.debug("Transactions cache HIT for user %s", request.user.id)
                return Response(cached)
            logger.debug("Transactions cache MISS for user %s", request.user.id)

        response = super().list(request, *args, **kwargs)

        if not has_filters:
            logger.debug("Transactions cache SET for user %s", request.user.id)
            cache.set(cache_key, response.data, CACHE_TTL)

        return response



@extend_schema_view(
    list=extend_schema(
        tags=["Internal Transfers"],
        summary="List internal transfers",
        description=(
            "Returns a paginated list of all internal (account-to-account) transfers for the authenticated user.\n\n"
            "**Filterable by:** `account`, `start_date`, `end_date`, `min_amount`, `max_amount`\n\n"
            "**Searchable by:** `note`, `from_account__account_name`, `to_account__account_name`\n\n"
            "**Sortable by:** `date`, `created_at`, `amount`"
        ),
        parameters=[
            OpenApiParameter("account", OpenApiTypes.INT, description="Filter by account ID (matches either from_account or to_account)."),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Filter transfers on or after this date (YYYY-MM-DD)."),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Filter transfers on or before this date (YYYY-MM-DD)."),
            OpenApiParameter("min_amount", OpenApiTypes.NUMBER, description="Filter transfers with amount ≥ this value."),
            OpenApiParameter("max_amount", OpenApiTypes.NUMBER, description="Filter transfers with amount ≤ this value."),
            OpenApiParameter("search", OpenApiTypes.STR, description="Search by note text."),
            OpenApiParameter("ordering", OpenApiTypes.STR, description="Sort field. Prefix with `-` for descending. Options: `date`, `created_at`, `amount`."),
        ],
        responses={200: InternalTransactionSerializer(many=True)},
    ),
    create=extend_schema(
        tags=["Internal Transfers"],
        summary="Create an internal transfer",
        description=(
            "Transfers money between two of the authenticated user's own accounts.\n\n"
            "**Required fields:** `from_account`, `to_account`, `amount`\n\n"
            "**Optional fields:** `note`, `date`\n\n"
            "**Validation rules:**\n"
            "- `from_account` and `to_account` must be different accounts.\n"
            "- `amount` must be greater than 0.\n"
            "- `from_account` must have sufficient balance.\n\n"
            "**Side effects:**\n"
            "- Deducts `amount` from `from_account.balance`\n"
            "- Adds `amount` to `to_account.balance`\n"
            "- Automatically creates a linked `Transaction` record for unified activity view"
        ),
        request=InternalTransactionSerializer,
        responses={
            201: InternalTransactionSerializer,
            400: OpenApiResponse(
                description="Validation error.",
                examples=[
                    OpenApiExample("Same account", value={"non_field_errors": ["From account and to account cannot be the same."]}),
                    OpenApiExample("Insufficient balance", value={"non_field_errors": ["From account balance is insufficient."]}),
                    OpenApiExample("Zero amount", value={"non_field_errors": ["Amount must be greater than 0."]}),
                ],
            ),
        },
        examples=[
            OpenApiExample(
                "Transfer between accounts",
                request_only=True,
                value={
                    "from_account": 1,
                    "to_account": 2,
                    "amount": "1000.00",
                    "date": "2024-01-15T14:00:00Z",
                    "note": "Moving savings to checking",
                },
            )
        ],
    ),
    retrieve=extend_schema(
        tags=["Internal Transfers"],
        summary="Retrieve an internal transfer",
        description="Returns the full details of a specific internal transfer.",
        responses={200: InternalTransactionSerializer},
    ),
)
class InternalTransactionViewSet(mixins.CreateModelMixin,
                                 mixins.ListModelMixin,
                                 mixins.RetrieveModelMixin,
                                 viewsets.GenericViewSet):
    """
    Create and Read for internal transfers.

    list     GET    /api/internal-transactions/
    create   POST   /api/internal-transactions/
    retrieve GET    /api/internal-transactions/{id}/
    """
    serializer_class = InternalTransactionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = TransactionResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = InternalTransactionFilter
    search_fields = ['note', 'from_account__account_name', 'to_account__account_name']
    ordering_fields = ['date', 'created_at', 'amount']

    def get_queryset(self):
        return InternalTransaction.objects.filter(
            user=self.request.user
        ).select_related('from_account', 'to_account').order_by('-date', '-created_at')

    @transaction.atomic
    def perform_create(self, serializer):
        from django.db.models import F
        from tracker.models import Transaction, Account
        
        instance = serializer.save(user=self.request.user)
        
        # Update balances
        Account.objects.filter(id=instance.from_account.id).update(balance=F('balance') - instance.amount)
        Account.objects.filter(id=instance.to_account.id).update(balance=F('balance') + instance.amount)
        
        # Create corresponding Transaction record
        tx = Transaction.objects.create(
            user=self.request.user,
            internal_transaction=instance,
            date=instance.date
        )
        tx.created_at = instance.created_at
        tx.save(update_fields=['created_at'])
