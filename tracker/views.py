from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
from .models import Account, Loan, Transaction, Contact, ContactAccount, TransactionSplit
from .serializers import AccountSerializer, LoanSerializer, TransactionSerializer, UserSerializer, ContactSerializer, ContactAccountSerializer
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.conf import settings
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Email, To, Content


@extend_schema_view(
    list=extend_schema(description='List all users', tags=['Users']),
    create=extend_schema(description='Register a new user', tags=['Users']),
    retrieve=extend_schema(description='Get user details', tags=['Users']),
    update=extend_schema(description='Update user', tags=['Users']),
    partial_update=extend_schema(description='Partially update user', tags=['Users']),
    destroy=extend_schema(description='Delete user', tags=['Users']),
)
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.AllowAny]

    def get_permissions(self):
        if self.action in ['create', 'forgot_password', 'reset_password']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        username = request.data.get('username')
        if User.objects.filter(username=username).exists():
            first_name = request.data.get('first_name', '').lower()
            last_name = request.data.get('last_name', '').lower()
            
            import random
            import string
            
            base = f"{first_name}{last_name}"
            if not base:
                base = "user"
            
            suggestion = ""
            while True:
                suffix = ''.join(random.choices(string.digits, k=3))
                suggestion = f"{base}{suffix}"
                if not User.objects.filter(username=suggestion).exists():
                    break
            
            return Response({
                'username': ['A user with that username already exists.'],
                'suggestion': suggestion
            }, status=status.HTTP_400_BAD_REQUEST)
            
        return super().create(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def forgot_password(self, request):
        email = request.data.get('email')
        if not email:
            return Response(
                {"detail": "Email is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            # Handle multiple users with same email by taking the first one
            user = User.objects.filter(email=email).first()
            if not user:
                return Response({'detail': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
                
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')
            reset_url = f"{frontend_url}/reset-password?uid={uid}&token={token}"

            subject = "Reset Your XPENSE Password"
            text_content = f"Click the link below to reset your password:\n\n{reset_url}"
            html_content = f"""
                                <p>Hello,</p>
                                <p>Click the button below to reset your password:</p>
                                <p>
                                    <a href="{reset_url}" style="padding:10px 15px;background:#4f46e5;color:white;text-decoration:none;border-radius:5px;">
                                        Reset Password
                                    </a>
                                </p>
                                <p>If you didn’t request this, you can ignore this email.</p>
                            """
            
            email_message = Mail(
                from_email=settings.DEFAULT_FROM_EMAIL,
                to_emails=email,
                subject=subject,
                html_content=html_content
            )
            sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
            sg.client.mail.send.post(request_body=email_message.get())

            return Response({"detail": "Password reset link sent to your email."})
            
            # send_mail(
            #     subject,
            #     text_content,
            #     html_content,
            #     getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@xpense.com'),
            #     [email],
            #     fail_silently=False,
            # )
            # return Response({'detail': 'Password reset link sent to your email.'})
        except Exception as e:
            # Log the error in production
            print(e)
            return Response({'detail': 'An error occurred. Please try again later.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def reset_password(self, request):
        uidb64 = request.data.get('uid')
        token = request.data.get('token')
        new_password = request.data.get('new_password')

        if not all([uidb64, token, new_password]):
            return Response(
                {"detail": "Invalid request."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response(
                {"detail": "Invalid or expired reset link."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not default_token_generator.check_token(user, token):
            return Response(
                {"detail": "Invalid or expired reset link."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate password strength
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response(
                {"detail": e.messages},
                status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()

        return Response({"detail": "Password has been reset successfully."})

@extend_schema_view(
    list=extend_schema(description='List all accounts for the authenticated user', tags=['Accounts']),
    create=extend_schema(description='Create a new account', tags=['Accounts']),
    retrieve=extend_schema(description='Get account details', tags=['Accounts']),
    update=extend_schema(description='Update account', tags=['Accounts']),
    partial_update=extend_schema(description='Partially update account', tags=['Accounts']),
    destroy=extend_schema(description='Delete account', tags=['Accounts']),
)
class AccountViewSet(viewsets.ModelViewSet):
    serializer_class = AccountSerializer

    def get_queryset(self):
        return Account.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

@extend_schema_view(
    list=extend_schema(description='List all loans for the authenticated user', tags=['Loans']),
    create=extend_schema(description='Create a new loan record', tags=['Loans']),
    retrieve=extend_schema(description='Get loan details', tags=['Loans']),
    update=extend_schema(description='Update loan', tags=['Loans']),
    partial_update=extend_schema(description='Partially update loan', tags=['Loans']),
    destroy=extend_schema(description='Delete loan', tags=['Loans']),
)
class LoanViewSet(viewsets.ModelViewSet):
    serializer_class = LoanSerializer

    def get_queryset(self):
        return Loan.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

@extend_schema_view(
    list=extend_schema(description='List all contacts for the authenticated user', tags=['Contacts']),
    create=extend_schema(description='Create a new contact', tags=['Contacts']),
    retrieve=extend_schema(description='Get contact details', tags=['Contacts']),
    update=extend_schema(description='Update contact', tags=['Contacts']),
    partial_update=extend_schema(description='Partially update contact', tags=['Contacts']),
    destroy=extend_schema(description='Delete contact', tags=['Contacts']),
)
class ContactViewSet(viewsets.ModelViewSet):
    serializer_class = ContactSerializer

    def get_queryset(self):
        return Contact.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class ContactAccountViewSet(viewsets.ModelViewSet):
    serializer_class = ContactAccountSerializer

    def get_queryset(self):
        return ContactAccount.objects.filter(contact__user=self.request.user)

@extend_schema_view(
    list=extend_schema(description='List all transactions for the authenticated user', tags=['Transactions']),
    create=extend_schema(description='Create a new transaction (automatically updates account balance and loan amounts)', tags=['Transactions']),
    retrieve=extend_schema(description='Get transaction details', tags=['Transactions']),
    update=extend_schema(description='Update transaction', tags=['Transactions']),
    partial_update=extend_schema(description='Partially update transaction', tags=['Transactions']),
    destroy=extend_schema(description='Delete transaction (reverses account and loan updates)', tags=['Transactions']),
)
class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer

    def get_queryset(self):
        from django.utils import timezone
        from datetime import timedelta
        
        queryset = Transaction.objects.filter(user=self.request.user)
        
        # Check for date filters
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            # Add time to end_date to include the whole day
            queryset = queryset.filter(date__lte=f"{end_date} 23:59:59")
            
        # Default to past 7 days only if it's list action and NO date filters provided
        if self.action == 'list' and not self.request.query_params.get('all') and not (start_date or end_date):
            seven_days_ago = timezone.now() - timedelta(days=7)
            queryset = queryset.filter(date__gte=seven_days_ago)
            
        return queryset.order_by('-date')

    @action(detail=False, methods=['get'])
    def download_excel(self, request):
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from django.http import HttpResponse
        from io import BytesIO
        from PIL import Image as PILImage
        import os
        from django.conf import settings
        
        # Get all transactions for export (respect date filters if any)
        transactions = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Set headers per user request
        headers = ['Date', 'Type', 'Amount', 'From Account', 'To Account', 'Contact', 'Note', 'Image']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 25
        
        for idx, t in enumerate(transactions, start=2):
            contact_name = f"{t.contact.first_name} {t.contact.last_name}" if t.contact else "-"
            
            # Format: Bank Name - Account Number - Account Owner
            def format_account(acc):
                if not acc: return "-"
                if acc.account_number:
                    return f"{acc.bank_name} - {acc.account_number} - {acc.account_name}"
                else:
                    return f"{acc.bank_name} - {acc.account_name}"

            from_acc_str = "-"
            to_acc_str = "-"
            
            if t.type == 'TRANSFER':
                from_acc_str = format_account(t.account)
                if t.to_account:
                    to_acc_str = format_account(t.to_account)
                elif t.to_contact_account:
                    if t.to_contact_account.account_number:
                        to_acc_str = f"{t.to_contact_account.account_name} - {t.to_contact_account.account_number} - {contact_name}"
                    else:
                        to_acc_str = f"{t.to_contact_account.account_name} - {contact_name}"
            elif t.splits.exists():
                split_accs = ", ".join([f"{format_account(s.account)} (Rs. {s.amount})" for s in t.splits.all()])
                if t.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                    # Money landed in multiple accounts
                    to_acc_str = split_accs
                    from_acc_str = contact_name
                else:
                    # Money spent from multiple accounts
                    from_acc_str = split_accs
                    to_acc_str = contact_name
            elif t.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                to_acc_str = format_account(t.account)
                from_acc_str = contact_name
            else: # EXPENSE, REPAYMENT, MONEY_LENT
                from_acc_str = format_account(t.account)
                to_acc_str = contact_name
            
            # Format date
            date_str = t.date.strftime('%Y-%m-%d %H:%M:%S') if t.date else "-"
            
            ws.append([
                date_str,
                t.get_type_display(),
                float(t.amount),
                from_acc_str,
                to_acc_str,
                contact_name,
                t.note,
                "" # Image placeholder
            ])
            
            # Add image if it exists
            if t.image:
                try:
                    # Use .open() which works for both local and remote (Cloudinary) storage
                    img_file = t.image.open()
                    pil_img = PILImage.open(img_file)
                    
                    # Resize maintaining aspect ratio
                    pil_img.thumbnail((150, 150))
                    
                    img_io = BytesIO()
                    # Ensure compatibility (convert to RGB if RGBA)
                    if pil_img.mode in ("RGBA", "P"):
                        pil_img = pil_img.convert("RGB")
                    pil_img.save(img_io, format='PNG')
                    img_io.seek(0)
                    
                    xl_img = OpenpyxlImage(img_io)
                    # Center the image in the cell loosely
                    ws.add_image(xl_img, f'H{idx}')
                    # Adjust row height to fit image
                    ws.row_dimensions[idx].height = 120
                    img_file.close() # Good practice to close
                except Exception as e:
                    ws[f'H{idx}'] = f"Error: {str(e)}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transactions_export.xlsx"'
        return response

    @transaction.atomic
    def perform_create(self, serializer):
        import json
        instance = serializer.save(user=self.request.user)
        
        # 1. Handle Automatic Loan Creation (if applicable)
        if instance.type in ['LOAN_TAKEN', 'MONEY_LENT'] and instance.contact and not instance.loan:
            loan_type = 'TAKEN' if instance.type == 'LOAN_TAKEN' else 'LENT'
            new_loan = Loan.objects.create(
                user=instance.user,
                contact=instance.contact,
                type=loan_type,
                total_amount=0, 
                remaining_amount=0,
                description=instance.note
            )
            instance.loan = new_loan
            instance.save()

        # 2. Update Account Balances
        from django.db.models import F
        splits_data = self.request.data.get('splits')
        if splits_data:
            import json
            from decimal import Decimal
            splits = json.loads(splits_data)
            for split in splits:
                acc_id = split['account']
                amt = Decimal(split['amount'])
                TransactionSplit.objects.create(transaction=instance, account_id=acc_id, amount=amt)
                
                # Update balance using F expression
                acc = Account.objects.filter(id=acc_id, user=self.request.user)
                if instance.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                    acc.update(balance=F('balance') + amt)
                else:
                    acc.update(balance=F('balance') - amt)
        elif instance.type == 'TRANSFER':
            # Handle Transfer logic
            if instance.account:
                Account.objects.filter(id=instance.account.id).update(balance=F('balance') - instance.amount)
            
            if instance.to_account:
                Account.objects.filter(id=instance.to_account.id).update(balance=F('balance') + instance.amount)
        elif instance.account:
            # Simple transaction (one account)
            acc = Account.objects.filter(id=instance.account.id)
            if instance.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                acc.update(balance=F('balance') + instance.amount)
            else:
                acc.update(balance=F('balance') - instance.amount)

        # 3. Update Loan Totals
        loan = instance.loan
        if loan:
            if instance.type == 'REPAYMENT':
                loan.remaining_amount -= instance.amount
            elif instance.type == 'REIMBURSEMENT':
                loan.remaining_amount -= instance.amount
            elif instance.type == 'LOAN_TAKEN':
                loan.remaining_amount += instance.amount
                loan.total_amount += instance.amount
            elif instance.type == 'MONEY_LENT':
                loan.remaining_amount += instance.amount
                loan.total_amount += instance.amount
            
            loan.is_closed = loan.remaining_amount <= 0
            loan.save()

    @transaction.atomic
    def perform_destroy(self, instance):
        from django.db.models import F
        # Reverse balance updates for all accounts involved
        splits = instance.splits.all()
        if splits.exists():
            for split in splits:
                acc = Account.objects.filter(id=split.account.id)
                if instance.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                    acc.update(balance=F('balance') - split.amount)
                else:
                    acc.update(balance=F('balance') + split.amount)
        elif instance.type == 'TRANSFER':
            if instance.account:
                Account.objects.filter(id=instance.account.id).update(balance=F('balance') + instance.amount)
            if instance.to_account:
                Account.objects.filter(id=instance.to_account.id).update(balance=F('balance') - instance.amount)
        elif instance.account:
            acc = Account.objects.filter(id=instance.account.id)
            if instance.type in ['INCOME', 'LOAN_TAKEN', 'REIMBURSEMENT']:
                acc.update(balance=F('balance') - instance.amount)
            else:
                acc.update(balance=F('balance') + instance.amount)

        # Reverse loan update
        loan = instance.loan
        if loan:
            if instance.type == 'REPAYMENT':
                loan.remaining_amount += instance.amount
            elif instance.type == 'REIMBURSEMENT':
                loan.remaining_amount += instance.amount
            elif instance.type == 'LOAN_TAKEN':
                loan.remaining_amount -= instance.amount
                loan.total_amount -= instance.amount
            elif instance.type == 'MONEY_LENT':
                loan.remaining_amount -= instance.amount
                loan.total_amount -= instance.amount
            
            loan.is_closed = loan.remaining_amount <= 0
            loan.save()
        
        instance.delete()
